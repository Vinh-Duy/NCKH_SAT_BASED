import csv
import time
from pathlib import Path

import networkx as nx


def log(message=""):
    print(message)
    with open("benchmark_run.log", "a", encoding="utf-8") as f:
        f.write(str(message) + "\n")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pysat.solvers import Glucose3


class OrderVars:
    def __init__(self, n_vertices, s):
        self.s = s
        self.n_vertices = n_vertices
        self.next_var = 1
        self.x = {}
        for v in range(n_vertices):
            self.x[v] = {}
            for i in range(s):
                self.x[v][i] = self.next_var
                self.next_var += 1

    def leq(self, v, i):
        if i < 0 or i >= self.s:
            return None
        return self.x[v][i]


def monotone_clauses(ov):
    clauses = []
    for v in range(ov.n_vertices):
        for i in range(ov.s - 1):
            clauses.append([-ov.x[v][i], ov.x[v][i + 1]])
    return clauses


def not_eq_literals(ov, v, a):
    lits = []
    l_a = ov.leq(v, a)
    if l_a is not None:
        lits.append(-l_a)
    l_am1 = ov.leq(v, a - 1)
    if l_am1 is not None:
        lits.append(l_am1)
    return lits


def forbid_close_labels(ov, u, v, t):
    clauses = []
    if t <= 0:
        return clauses
    for a in range(ov.s + 1):
        lo = max(0, a - t + 1)
        hi = min(ov.s, a + t - 1)
        for b in range(lo, hi + 1):
            clauses.append(not_eq_literals(ov, u, a) + not_eq_literals(ov, v, b))
    return clauses


def solve_lhk(n_vertices, edges, dist2_pairs, h, k, s):
    ov = OrderVars(n_vertices, s)
    cnf = []
    cnf += monotone_clauses(ov)
    for u, v in edges:
        cnf += forbid_close_labels(ov, u, v, h)
    for u, v in dist2_pairs:
        cnf += forbid_close_labels(ov, u, v, k)

    if any(len(c) == 0 for c in cnf):
        return None, None
    return cnf, ov


def get_graph_data(G):
    edges = list(G.edges())
    dist2_pairs = []
    nodes = list(G.nodes())

    for i in range(len(nodes)):
        for j in range(i + 1, len(nodes)):
            u, v = nodes[i], nodes[j]
            try:
                if nx.shortest_path_length(G, source=u, target=v) == 2:
                    dist2_pairs.append((u, v))
            except nx.NetworkXNoPath:
                pass
    return edges, dist2_pairs


def compute_lower_bound(G, h=2, k=1):
    if G.number_of_nodes() == 0:
        return 0
    max_degree = max(dict(G.degree()).values(), default=0)
    return max(0, max_degree + h - 1)


def estimate_upper_bound(G, h=2, k=1):
    """Upper bound = |V| + (max_degree)^2 theo gợi ý từ cô."""
    n = G.number_of_nodes()
    if n == 0:
        return 0

    max_degree = max(dict(G.degree()).values(), default=0)
    return n + (max_degree ** 2)


def run_benchmark(graph_name, G, n, h=2, k=1, max_span=None, timeout_sec=60):
    """Chạy benchmark từ cao xuống thấp (descending), có timeout.
    
    Args:
        timeout_sec: timeout tính bằng giây (30-50s cho lần đầu, 600-900s cho retry)
    """
    edges, dist2_pairs = get_graph_data(G)
    lower_bound = compute_lower_bound(G, h=h, k=k)
    if max_span is None:
        max_span = estimate_upper_bound(G, h=h, k=k)

    if max_span < lower_bound:
        max_span = lower_bound

    start_time = time.time()
    # Tìm kiếm từ cao xuống thấp (descending)
    for s in range(max_span, lower_bound - 1, -1):
        # Kiểm tra timeout
        if time.time() - start_time > timeout_sec:
            return {
                "Graph": graph_name,
                "n": n,
                "var": None,
                "clause": None,
                "time": round(time.time() - start_time, 6),
                "lambda": None,
                "status": "TIMEOUT",
            }

        cnf, ov = solve_lhk(n, edges, dist2_pairs, h, k, s)
        if cnf is None:
            continue

        with Glucose3() as solver:
            for clause in cnf:
                solver.add_clause(clause)

            if solver.solve():
                end_time = time.time()
                time_taken = round(end_time - start_time, 6)
                num_vars = ov.next_var - 1
                num_clauses = len(cnf)
                return {
                    "Graph": graph_name,
                    "n": n,
                    "var": num_vars,
                    "clause": num_clauses,
                    "time": time_taken,
                    "lambda": s,
                    "status": "OPT",
                }

    return {
        "Graph": graph_name,
        "n": n,
        "var": None,
        "clause": None,
        "time": round(time.time() - start_time, 6),
        "lambda": None,
        "status": "UNSOLVED",
    }


def export_excel(results, csv_filename="ket_qua_SAT.csv", excel_filename="ket_qua_SAT.xlsx"):
    keys = ["Graph", "n", "var", "clause", "time", "lambda", "status"]

    csv_path = Path(csv_filename)
    with open(csv_path, mode="w", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=keys)
        writer.writeheader()
        writer.writerows(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "SAT Benchmark"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(keys)
    for row in results:
        ws.append([row.get(k) for k in keys])

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(keys)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    wb.save(excel_filename)
    return csv_path, Path(excel_filename)


def export_separate_files(c_results, k_results, q_results):
    """Export results into separate files for each graph family."""
    keys = ["Graph", "n", "var", "clause", "time", "lambda", "status"]
    
    def create_file(results, prefix):
        csv_filename = f"ket_qua_{prefix}.csv"
        excel_filename = f"ket_qua_{prefix}.xlsx"
        
        csv_path = Path(csv_filename)
        with open(csv_path, mode="w", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=keys)
            writer.writeheader()
            writer.writerows(results)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"SAT {prefix}"
        
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws.append(keys)
        for row in results:
            ws.append([row.get(k) for k in keys])
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(keys)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        
        wb.save(excel_filename)
        return csv_path, Path(excel_filename)
    
    c_csv, c_excel = create_file(c_results, "C")
    k_csv, k_excel = create_file(k_results, "K")
    q_csv, q_excel = create_file(q_results, "Q")
    
    return (c_csv, c_excel, k_csv, k_excel, q_csv, q_excel)


def analyze_pattern(results):
    """Phân tích pattern lambda theo n, đặc biệt cho n>5 và n>11."""
    log("\n=== PHÂN TÍCH PATTERN ===")
    by_graph_type = {}
    for r in results:
        if r["lambda"] is not None:
            graph_type = r["Graph"].split("_")[0]
            if graph_type not in by_graph_type:
                by_graph_type[graph_type] = []
            by_graph_type[graph_type].append((r["n"], r["lambda"]))
    
    for graph_type in sorted(by_graph_type.keys()):
        pairs = sorted(by_graph_type[graph_type], key=lambda x: x[0])
        log(f"\n{graph_type}:")
        small = [p for p in pairs if p[0] <= 5]
        medium = [p for p in pairs if 5 < p[0] <= 11]
        large = [p for p in pairs if p[0] > 11]
        if small:
            log(f"  n <= 5: {small}")
        if medium:
            log(f"  5 < n <= 11: {medium}")
        if large:
            log(f"  n > 11: {large}")
            if len(large) >= 2:
                diffs = [large[i+1][1] - large[i][1] for i in range(len(large)-1)]
                log(f"    Chênh lệch lambda: {diffs}")


def main():
    Path("benchmark_run.log").write_text("", encoding="utf-8")
    c_results = []
    k_results = []
    q_results = []

    log("Chạy benchmark SAT cho các đồ thị chuẩn (n lên đến 50)...")
    log("Lưu ý: Có thể mất vài lúc vì n lớn.\n")

    # Cycle graphs C_n từ n=3 đến n=50
    log("--- Đồ thị chu trình C_n (n=3..50) ---")
    for n in range(3, 51):
        G = nx.cycle_graph(n)
        max_span = estimate_upper_bound(G, h=2, k=1)
        res = run_benchmark(f"C_{n}", G, n, h=2, k=1, max_span=max_span, timeout_sec=45)
        c_results.append(res)
        if n % 5 == 0 or n <= 10:
            log(f"Hoàn thành C_{n} (lambda={res['lambda']})")

    # Complete graphs K_n từ n=3 đến n=12
    log("\n--- Đồ thị đầy đủ K_n (n=3..12) ---")
    for n in range(3, 13):
        G = nx.complete_graph(n)
        max_span = estimate_upper_bound(G, h=2, k=1)
        res = run_benchmark(f"K_{n}", G, n, h=2, k=1, max_span=max_span, timeout_sec=120)
        k_results.append(res)
        if n % 5 == 0 or n <= 10:
            log(f"Hoàn thành K_{n} (lambda={res['lambda']})")

    # Hypercube Q_n từ n=2 đến n=6
    log("\n--- Đồ thị siêu khối Q_n (n=2..6) ---")
    for n in range(2, 7):
        G = nx.hypercube_graph(n)
        num_nodes = 2 ** n
        G = nx.convert_node_labels_to_integers(G)
        max_span = estimate_upper_bound(G, h=2, k=1)
        res = run_benchmark(f"Q_{n}", G, num_nodes, h=2, k=1, max_span=max_span, timeout_sec=60)
        q_results.append(res)
        log(f"Hoàn thành Q_{n} (lambda={res['lambda']})")

    log("\n=== PHÂN TÍCH PATTERN ===")
    log("\n--- Đồ thị chu trình C_n ---")
    analyze_pattern(c_results)
    log("\n--- Đồ thị đầy đủ K_n ---")
    analyze_pattern(k_results)
    log("\n--- Đồ thị siêu khối Q_n ---")
    analyze_pattern(q_results)
    
    c_csv, c_excel, k_csv, k_excel, q_csv, q_excel = export_separate_files(c_results, k_results, q_results)
    log(f"\nXONG! Cycle graphs: {c_csv} & {c_excel}")
    log(f"XONG! Complete graphs: {k_csv} & {k_excel}")
    log(f"XONG! Hypercube graphs: {q_csv} & {q_excel}")


if __name__ == "__main__":
    main()
